from openpyxl import Workbook, load_workbook
from pprint import pprint

wb = load_workbook("inventory.xlsx")
sheet = wb['Sheet1']

"""
Genero una list pulita dei dati
da parsare
"""
def get_data(file_sheet) : 

    data = []

    i = int(0);
    for row in sheet.iter_rows() :
        this_row = []
        i += 1
        n = int(0);
        for cell in row :
            if cell.value != None and i > 1 and n > 0:
                if type( cell.value ) == float :
                    value = int( cell.value )
                else :
                    value = cell.value
                this_row.append(  value  )
            n += 1
        if this_row : 
            data.append( this_row )
    return data;


"""
Filtro i dati per azienda
"""
def filter_by_org( company, data ) :
    return list( filter( lambda item : item[2] == company, data ) )


"""
Ottengo i totali per ogni azienda.
Posso decidere se ottenere il totale dei prodotti o dei prezzi,
cambiando l'indice
"""
def total_by_org( index, company, data ) :
    data = filter_by_org( company, data )
    total = int(0)
    for row in data :
        total += row[ index ]
    return total

"""
Ottengo tutti i prodotti minori di 10
"""
def filter_less_10( data ) :
    return list( filter( lambda item : item[0] < 10, data ) )

"""
Calcolo il costo totale riga per riga di ogni prodotto
e salvo in una nuova colonna dell'excel
"""
def get_total( data ) :
    sheet.cell( 1 , 5 ).value = 'Sum'
    for row in range(2, data.max_row ) :
        inventory = sheet.cell( row , column = 2 ).value
        price = sheet.cell( row , column = 3 ).value
        total = inventory * price
        sheet.cell( row, 5 ).value = total
        wb.save('new-inventory.xlsx')


# get_total( sheet )

# pprint( filter_less_10( get_data(sheet) ) )
# pprint( total_by_org(0, 'AAA Company', get_data(sheet) ) )
# pprint( total_by_org(0, 'BBB Company', get_data(sheet) ) )
# pprint( total_by_org(0, 'CCC Company', get_data(sheet) ) )